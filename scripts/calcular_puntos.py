"""
Porra Mundial 2026 - Calculador de puntos
==========================================
Sistema de puntos completo según las reglas oficiales:
 
FASE DE GRUPOS:
  - Resultado exacto: 4 pts
  - Signo correcto (G/E/P): 1 pt
  - Clasificados 1º y 2º de grupo: +2 pts por acierto
 
ELIMINATORIAS:
  - Resultado exacto (90 min): 4 pts
    · Si el partido acabó en penaltis, además hay que acertar
      quién pasa de ronda para llevarse los 4 pts completos.
      Si solo se acierta el empate pero no quién pasa: 1 pt.
  - Signo correcto: 1 pt
  - Acertar quién pasa de cada ronda (cuadro de clasificados): +3 pts
 
POSICIÓN FINAL:
  - Campeón: +10 pts
  - Subcampeón: +5 pts
  - 3er puesto: +2 pts
  - 4º puesto: +2 pts
 
PREMIOS ESPECIALES (5 pts cada uno):
  - Campeón, Subcampeón, 3er puesto
  - Bota de Oro/Plata/Bronce
  - Balón de Oro/Plata/Bronce
"""
 
import json
import re
import glob
from pathlib import Path
from datetime import datetime
import openpyxl
 
# ─── RUTAS ───────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent
EXCELS_DIR = BASE_DIR / "excels"
print(EXCELS_DIR)
RESULTS_FILE = BASE_DIR / "resultados_reales.json"
OUTPUT_FILE  = BASE_DIR / "data.json"
 
# ─── PUNTUACIÓN ──────────────────────────────────────────────────────────────
PTS_EXACTO        = 4   # marcador exacto (incluye el punto de signo)
PTS_SIGNO         = 1   # solo signo correcto
PTS_CLASIFICADO   = 2   # acertar 1º o 2º de grupo
PTS_ELIMINATORIA  = 3   # acertar quién pasa de ronda
PTS_CAMPEON       = 10
PTS_SUBCAMPEON    = 5
PTS_TERCERO       = 2
PTS_CUARTO        = 2
PTS_ESPECIAL      = 5   # botas y balones

# Etiquetas de clasificados por ronda, en la hoja Pool
ETIQUETAS_CLASIFICADOS = [
    "Dieciseisavofinalista", "Octavofinalista", "Cuartofinalista",
    "Semifinalista", "Finalista", "3º y 4º puesto",
    "🥇Campeón", "🥈Subcampeón", "🥉3º puesto",
    "Bota de Oro", "Bota de Plata", "Bota de Bronce",
    "Balón de Oro", "Balón de Plata", "Balón de Bronce",
]
 
 
def sign(local, visitante):
    if local > visitante: return "1"
    if local < visitante: return "2"
    return "X"
 
 
def puntuar_partido(pronostico_str, resultado_str):
    """Puntúa un partido de fase de grupos. Devuelve None si no jugado."""
    if not resultado_str:
        return None
    m = re.match(r"([12X])\|(\d+)-(\d+)", str(pronostico_str).strip())
    if not m:
        return None
    p_signo = m.group(1)
    p_local, p_visit = int(m.group(2)), int(m.group(3))
 
    m2 = re.match(r"(\d+)-(\d+)", str(resultado_str).strip())
    if not m2:
        return None
    r_local, r_visit = int(m2.group(1)), int(m2.group(2))
    r_signo = sign(r_local, r_visit)
 
    if p_local == r_local and p_visit == r_visit:
        return PTS_EXACTO
    elif p_signo == r_signo:
        return PTS_SIGNO
    else:
        return 0


def puntuar_partido_elim(pronostico_str, resultado_str, ganador_real=None, equipo_predicho_pasa=None):
    """
    Puntúa un partido de eliminatorias.

    pronostico_str: 'Equipo1-Equipo2·signo|g1-g2' (de la hoja Pool, ENFRENTAMIENTOS)
    resultado_str:  'g1-g2' (resultado real en los 90 minutos)
    ganador_real:   nombre del equipo que realmente pasó de ronda (solo relevante
                    si el resultado real fue empate, es decir, se decidió en penaltis)
    equipo_predicho_pasa: nombre del equipo que el participante puso como clasificado
                          de esa ronda (ej. en 'Octavofinalista'). Solo se usa cuando
                          el resultado real terminó en empate.

    Reglas:
      - Si el marcador NO es exacto: 1 pt si el signo coincide, si no 0 pts.
      - Si el marcador SÍ es exacto y el resultado real NO fue empate: 4 pts.
      - Si el marcador SÍ es exacto y el resultado real SÍ fue empate (penaltis):
          · Si el participante acertó además quién pasa de ronda: 4 pts.
          · Si no lo acertó (o no tenemos ese dato): 1 pt.
    """
    if not resultado_str or not pronostico_str:
        return None
    p_str = str(pronostico_str).strip()
    if "·" in p_str:
        p_str = p_str.split("·", 1)[1]
    m = re.match(r"([12X])\|(\d+)-(\d+)", p_str)
    if not m:
        return None
    p_signo, p_local, p_visit = m.group(1), int(m.group(2)), int(m.group(3))

    m2 = re.match(r"(\d+)-(\d+)", str(resultado_str).strip())
    if not m2:
        return None
    r_local, r_visit = int(m2.group(1)), int(m2.group(2))
    r_signo = sign(r_local, r_visit)

    marcador_exacto = (p_local == r_local and p_visit == r_visit)

    if not marcador_exacto:
        return PTS_SIGNO if p_signo == r_signo else 0

    # Marcador exacto
    if r_signo != "X":
        # No hubo empate en el resultado real: no hay ambigüedad de quién pasa
        return PTS_EXACTO

    # El resultado real fue empate (se decidió en penaltis)
    if ganador_real and equipo_predicho_pasa and equipo_predicho_pasa.strip() == ganador_real.strip():
        return PTS_EXACTO
    else:
        return PTS_SIGNO
 
 
def leer_excel(excel_path):
    """Lee hoja Pool y devuelve (pronosticos_dict, nombre).

    El dict 'datos' incluye, además de los pronósticos sueltos, listas
    acumuladas de clasificados bajo la clave '<Etiqueta>_LISTA' (por ejemplo
    'Octavofinalista_LISTA') porque cada etiqueta se repite muchas veces en
    la hoja (una fila por equipo clasificado).
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  ERROR abriendo {Path(excel_path).name}: {e}")
        return {}, Path(excel_path).stem
 
    datos = {}
    nombre = "DESCONOCIDO"
 
    # ── Nombre desde Pool ──
    if "Pool" in wb.sheetnames:
        ws = wb["Pool"]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            b = str(row[1]).strip() if row[1] is not None else ""
            c = str(row[2]).strip() if row[2] is not None else ""
 
            if b == "Nombre" and c:
                nombre = c
 
            # Partidos de grupos: solo si NO es eliminatoria
            if b and c and "|" in c and "-" in b and "·" not in c:
                datos[b] = c

            # Partidos eliminatorias
            if b and c and "·" in c and "|" in c:
                nombre_partido = c.split("·", 1)[0].strip()
                datos[nombre_partido] = c
 
            # Posiciones de grupo: col B = "1º GRUPO A", col C = "México"
            if b and c and "GRUPO" in b.upper() and "|" not in c:
                datos[b] = c
 
            # Clasificados eliminatorias / cuadro honor
            # col B = etiqueta, col C = equipo
            # NOTA: cada etiqueta se repite muchas veces (una fila por equipo
            # clasificado), por eso acumulamos en listas además de guardar
            # el último valor (compatibilidad con el código anterior).
            for label in ETIQUETAS_CLASIFICADOS:
                if label in b and c:
                    datos[b] = c
                    datos.setdefault(label + "_LISTA", []).append(c)
 
    wb.close()
 
    if nombre == "DESCONOCIDO":
        nombre = Path(excel_path).stem.replace("Excel-Mundial-2026__", "").replace("_", " ")
 
    return datos, nombre
 
 
def cargar_resultados():
    if not RESULTS_FILE.exists():
        print(f"AVISO: No existe {RESULTS_FILE}")
        return {}
    with open(RESULTS_FILE, encoding="utf-8") as f:
        return json.load(f)


def encontrar_equipo_predicho(nombre_partido, lista_clasificados):
    """De los dos equipos que se enfrentan en 'nombre_partido' (formato
    'Equipo1-Equipo2'), devuelve cuál de los dos aparece en la lista de
    equipos que el participante marcó como clasificados de esa ronda.
    """
    if not lista_clasificados:
        return None
    equipos = nombre_partido.split("-")
    if len(equipos) != 2:
        return None
    lista_normalizada = [e.strip() for e in lista_clasificados]
    for eq in equipos:
        if eq.strip() in lista_normalizada:
            return eq.strip()
    return None
 
 
def calcular_clasificacion():
    print("=" * 60)
    print("  PORRA MUNDIAL 2026 - Calculando clasificación")
    print("=" * 60)
 
    reales = cargar_resultados()
 
    # Separar por tipo
    partidos_grupos    = {k: v for k, v in reales.items() if isinstance(v, str) and re.match(r"\d+-\d+", v)}
    clasificados_grupo = reales.get("clasificados_grupo", {})
    eliminatorias      = reales.get("eliminatorias", {})
    posicion_final     = reales.get("posicion_final", {})
    premios_especiales = reales.get("premios_especiales", {})
    ganadores_penaltis = reales.get("ganadores_penaltis", {})
 
    print(f"\n  Partidos grupos con resultado: {len(partidos_grupos)}")
 
    excels = sorted(glob.glob(str(EXCELS_DIR / "*.xlsx")))
    print(f"  Participantes: {len(excels)}\n")
 
    participantes = []
 
    for excel_path in excels:
        print(f"  Procesando: {Path(excel_path).name}")
        datos, nombre = leer_excel(excel_path)
        print(f"    Nombre: {nombre}  |  Datos leídos: {len(datos)}")
 
        pts_total   = 0
        exactos     = 0
        signos      = 0
        fallos      = 0
        n_partidos  = 0
        detalle     = []
        detalle_extra = []
 
        # ── 1. Partidos de fase de grupos ──
        for partido, resultado_real in partidos_grupos.items():
            if not resultado_real:
                continue
            pronostico = datos.get(partido)
            if pronostico is None:
                continue
 
            pts = puntuar_partido(pronostico, resultado_real)
            if pts is None:
                continue
 
            n_partidos += 1
            pts_total  += pts
            if pts == PTS_EXACTO: exactos += 1
            elif pts == PTS_SIGNO: signos += 1
            else: fallos += 1
 
            detalle.append({
                "partido": partido,
                "pronostico": pronostico,
                "resultado": resultado_real,
                "puntos": pts
            })

        # ── Partidos de eliminatorias ──
        partidos_elim = reales.get("partidos_eliminatorias", {})
        pts_elim_partidos = 0

        for partido, resultado_real in partidos_elim.items():
            if not resultado_real:
                continue
            pronostico = datos.get(partido)
            if pronostico is None:
                continue

            ganador_real = ganadores_penaltis.get(partido)
            equipo_predicho_pasa = None
            if ganador_real:
                # Solo nos hace falta buscar el equipo predicho si el partido
                # se decidió por penaltis; buscamos en todas las listas de
                # clasificados (Dieciseisavofinalista, Octavofinalista, etc.)
                for label in ETIQUETAS_CLASIFICADOS:
                    lista = datos.get(label + "_LISTA")
                    equipo_predicho_pasa = encontrar_equipo_predicho(partido, lista)
                    if equipo_predicho_pasa:
                        break

            pts = puntuar_partido_elim(pronostico, resultado_real, ganador_real, equipo_predicho_pasa)
            if pts is None:
                continue
            n_partidos += 1
            pts_elim_partidos += pts
            if pts == PTS_EXACTO: exactos += 1
            elif pts == PTS_SIGNO: signos += 1
            else: fallos += 1
            detalle.append({
                "partido": partido,
                "pronostico": pronostico,
                "resultado": resultado_real,
                "puntos": pts,
                "fase": "eliminatoria"
            })

        pts_total += pts_elim_partidos
 
        # ── 2. Clasificados de grupo (1º y 2º) ──
        pts_clasificados = 0
        for pos_key, equipo_real in clasificados_grupo.items():
            if not equipo_real:
                continue
            pronostico = datos.get(pos_key)
            if pronostico and pronostico.strip() == equipo_real.strip():
                pts_clasificados += PTS_CLASIFICADO
                detalle_extra.append({"concepto": f"Clasificado: {pos_key}", "equipo": equipo_real, "puntos": PTS_CLASIFICADO})

        pts_total += pts_clasificados
 
        # ── 3. Eliminatorias (quién pasa de ronda) ──
        pts_elim = 0
        for ronda, equipos_reales in eliminatorias.items():
            if not equipos_reales:
                continue
            equipos_reales_norm = [e.strip() for e in equipos_reales]
            # Usar la lista acumulada de esa etiqueta (puede haber muchos equipos
            # bajo la misma clave, ej. 16 x "Dieciseisavofinalista")
            lista_predicha = datos.get(ronda + "_LISTA", [])
            for equipo_predicho in lista_predicha:
                if equipo_predicho.strip() in equipos_reales_norm:
                    pts_elim += PTS_ELIMINATORIA
                    detalle_extra.append({"concepto": ronda, "equipo": equipo_predicho, "puntos": PTS_ELIMINATORIA})
 
        pts_total += pts_elim
 
        # ── 4. Posición final ──
        pts_posicion = 0
        mapeo_pos = {
            "campeon":    ("🥇Campeón",    PTS_CAMPEON),
            "subcampeon": ("🥈Subcampeón", PTS_SUBCAMPEON),
            "tercero":    ("🥉3º puesto",  PTS_TERCERO),
        }
        for clave, (label, pts_val) in mapeo_pos.items():
            real = posicion_final.get(clave)
            if not real:
                continue
            pron = datos.get(label, "")
            if pron and pron.strip() == real.strip():
                pts_posicion += pts_val
                detalle_extra.append({"concepto": label, "equipo": real, "puntos": pts_val})
 
        pts_total += pts_posicion
 
        # ── 5. Premios especiales ──
        pts_premios = 0
        mapeo_premios = {
            "bota_oro":     "Bota de Oro",
            "bota_plata":   "Bota de Plata",
            "bota_bronce":  "Bota de Bronce",
            "balon_oro":    "Balón de Oro",
            "balon_plata":  "Balón de Plata",
            "balon_bronce": "Balón de Bronce",
        }
        for clave, label in mapeo_premios.items():
            real = premios_especiales.get(clave)
            if not real:
                continue
            pron = datos.get(label, "")
            if pron and pron.strip().lower() == real.strip().lower():
                pts_premios += PTS_ESPECIAL
                detalle_extra.append({"concepto": label, "jugador": real, "puntos": PTS_ESPECIAL})
 
        pts_total += pts_premios
 
        participantes.append({
            "nombre":       nombre,
            "puntos":       pts_total,
            "exactos":      exactos,
            "ganadores":    signos,
            "fallos":       fallos,
            "partidos":     n_partidos,
            "pts_clasificados": pts_clasificados,
            "pts_elim_partidos": pts_elim_partidos,
            "pts_elim":     pts_elim,
            "pts_posicion": pts_posicion,
            "pts_premios":  pts_premios,
            "detalle":      detalle,
            "detalle_extra": detalle_extra,
            "archivo":      Path(excel_path).name
        })
 
    participantes.sort(key=lambda x: (-x["puntos"], -x["exactos"]))
    for i, p in enumerate(participantes):
        p["posicion"] = i + 1
 
    output = {
        "actualizado":      datetime.now().strftime("%d/%m/%Y %H:%M"),
        "partidos_jugados": len(partidos_grupos),
        "participantes":    participantes
    }
 
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
 
    print(f"\n  ✓ data.json generado en: {OUTPUT_FILE}")
    print(f"\n{'─'*50}")
    print(f"  {'#':<4} {'NOMBRE':<20} {'PTS':<6} {'Exactos':<9} {'Signo':<7} {'Fallos'}")
    print(f"{'─'*50}")
    for p in participantes:
        print(f"  {p['posicion']:<4} {p['nombre']:<20} {p['puntos']:<6} {p['exactos']:<9} {p['ganadores']:<7} {p['fallos']}")
    print(f"{'─'*50}\n")
 
    return output
 
 
if __name__ == "__main__":
    calcular_clasificacion()